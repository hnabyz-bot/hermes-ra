#!/usr/bin/env python3
"""
Hermes NAS Scanner API v1
Issue #9: Phase 1-1 NAS 파일 변경 감지
- POST /scan  → NAS 스캔 + DB 비교 → 변경 목록 반환
- GET  /health → 상태 확인
"""
import json
import hashlib
import os
import re
import psycopg2
from http.server import BaseHTTPRequestHandler
from socketserver import ThreadingMixIn
from http.server import HTTPServer
from datetime import datetime

# 설정
NAS_TARGETS = [
    "/mnt/nas-ra/공통자료/RA/99_4. 한지민(241120~260508)/",
]
SCAN_EXTENSIONS = {".xlsx", ".pptx", ".pdf", ".docx"}
SKIP_PREFIX = ("~$", ".")

DB_CONFIG = {
    "host": os.environ.get("N8N_DB_HOST", "localhost"),
    "port": int(os.environ.get("N8N_DB_PORT", "5432")),
    "dbname": os.environ.get("N8N_DB_NAME", "n8n"),
    "user": os.environ.get("N8N_DB_USER", "n8n"),
    "password": os.environ.get("N8N_DB_PASSWORD", ""),
}

# OP WP 매핑 (파일명 키워드 → WP ID)
WP_FILE_MAP = [
    ("인수인계서", 638),
    ("해외 등록 대장", 638),
    ("인수인계 계획서", 638),
    ("인증현황", 638),
]


def md5_file(path: str) -> str:
    h = hashlib.md5()
    try:
        with open(path, "rb") as f:
            for chunk in iter(lambda: f.read(65536), b""):
                h.update(chunk)
        return h.hexdigest()
    except Exception:
        return ""


def scan_nas():
    """NAS 대상 경로의 파일을 스캔하여 {path, md5, mtime, size} 목록 반환"""
    results = []
    for base_dir in NAS_TARGETS:
        if not os.path.exists(base_dir):
            continue
        for fname in os.listdir(base_dir):
            if any(fname.startswith(p) for p in SKIP_PREFIX):
                continue
            ext = os.path.splitext(fname)[1].lower()
            if ext not in SCAN_EXTENSIONS:
                continue
            fpath = os.path.join(base_dir, fname)
            if not os.path.isfile(fpath):
                continue
            md5 = md5_file(fpath)
            if not md5:
                continue
            stat = os.stat(fpath)
            results.append({
                "path": fpath,
                "filename": fname,
                "md5": md5,
                "mtime": int(stat.st_mtime),
                "size": stat.st_size
            })
    return results


def get_wp_id(filename: str) -> int:
    for kw, wp_id in WP_FILE_MAP:
        if kw in filename:
            return wp_id
    return 638  # 기본값


def get_xlsx_sheet_summary(path: str) -> str:
    """xlsx 시트 목록 + 행 수 요약"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        parts = []
        for sn in wb.sheetnames:
            ws = wb[sn]
            cnt = sum(1 for row in ws.iter_rows(values_only=True) if any(c for c in row if c is not None))
            parts.append(f"{sn}({cnt}행)")
        wb.close()
        return ", ".join(parts)
    except Exception as e:
        return f"파싱오류: {str(e)[:60]}"


def detect_changes():
    """스캔 → DB 비교 → 변경사항 반환 + DB 업데이트"""
    scanned = scan_nas()
    changes = []

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # 기존 해시 로드
    cur.execute("SELECT path, md5 FROM nas_file_hashes")
    existing = {row[0]: row[1] for row in cur.fetchall()}

    for f in scanned:
        path = f["path"]
        prev_md5 = existing.get(path)

        if prev_md5 is None:
            change_type = "created"
        elif prev_md5 != f["md5"]:
            change_type = "modified"
        else:
            # 변경 없음 — last_checked만 업데이트
            cur.execute(
                "UPDATE nas_file_hashes SET last_checked=NOW() WHERE path=%s",
                (path,)
            )
            continue

        # 변경 있음
        xlsx_summary = ""
        if path.endswith(".xlsx"):
            xlsx_summary = get_xlsx_sheet_summary(path)

        wp_id = get_wp_id(f["filename"])
        change = {
            **f,
            "change_type": change_type,
            "old_md5": prev_md5,
            "wp_id": wp_id,
            "xlsx_summary": xlsx_summary,
        }
        changes.append(change)

        # nas_file_hashes upsert
        cur.execute(
            """INSERT INTO nas_file_hashes(path, md5, mtime, size_bytes, last_checked, last_changed)
               VALUES(%s,%s,%s,%s,NOW(),NOW())
               ON CONFLICT(path) DO UPDATE SET
                 md5=EXCLUDED.md5, mtime=EXCLUDED.mtime, size_bytes=EXCLUDED.size_bytes,
                 last_checked=NOW(), last_changed=NOW()""",
            (path, f["md5"], f["mtime"], f["size"])
        )
        # nas_change_log 기록
        cur.execute(
            "INSERT INTO nas_change_log(path, old_md5, new_md5, change_type) VALUES(%s,%s,%s,%s)",
            (path, prev_md5, f["md5"], change_type)
        )

    conn.commit()
    cur.close()
    conn.close()
    return changes


class Handler(BaseHTTPRequestHandler):
    def log_message(self, format, *args):
        pass

    def do_GET(self):
        if self.path == "/health":
            resp = json.dumps({"status": "ok", "service": "nas-scanner"}).encode()
            self.send_response(200)
            self.send_header("Content-Type", "application/json")
            self.send_header("Content-Length", len(resp))
            self.end_headers()
            self.wfile.write(resp)
        else:
            self.send_response(404)
            self.end_headers()

    def do_POST(self):
        if self.path == "/scan":
            try:
                changes = detect_changes()
                resp = json.dumps({
                    "changes": changes,
                    "count": len(changes),
                    "scanned_at": datetime.now().isoformat()
                }, ensure_ascii=False).encode("utf-8")
                self.send_response(200)
                self.send_header("Content-Type", "application/json; charset=utf-8")
                self.send_header("Content-Length", len(resp))
                self.end_headers()
                self.wfile.write(resp)
            except Exception as e:
                err = json.dumps({"error": str(e)}).encode()
                self.send_response(500)
                self.send_header("Content-Type", "application/json")
                self.end_headers()
                self.wfile.write(err)
        else:
            self.send_response(404)
            self.end_headers()


class ThreadingHTTPServer(ThreadingMixIn, HTTPServer):
    daemon_threads = True


if __name__ == "__main__":
    server = ThreadingHTTPServer(("0.0.0.0", 7789), Handler)
    print("Hermes NAS Scanner API v1 running on port 7789")
    server.serve_forever()
